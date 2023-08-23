


from openpyxl import Workbook
# from openpyxl.writer.excel import save_virtual_workbook
from io import BytesIO
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment,Font,Color
from openpyxl.styles.colors import BLUE



import os
import re
from datetime import date
from scopus.sc_forms import DataScForm
#import io


# PATH_LOAD_DEFAULT = ''
""""
    file_stream = io.BytesIO()
    file_stream.write(b"Hello, World!")

    # Возвращаемся в начало файла, чтобы Flask мог его прочитать
    file_stream.seek(0)

    # Используйте send_file для отправки данных клиенту как файл
    return send_file(
        file_stream,
        attachment_filename='example.txt',
        as_attachment=True,
        mimetype='text/plain'
    )
"""

#----------------------------------------------------------------
class ExportExcel():
   
    def __init__(self):
        self.mas_Doc=(4.0, 80.0, 10.0, 10, 55.0, 13.0,10.0)
        self.mas_note=(20.0, 12.0, 60.0, 12)
        self.mas_aut_=(4,45,10,20,80,19)
        self.title_note=('Кафедра','Сумма','ФИО','Цитирования')
        self.all_title=('#','Название статьи',"Год","Тип публ.",'Авторы/Автор','Кафедра','Цитир.')
        self.title_aut_=('#','Прізвище, ім’я, по батькові працівника ЗВО','Кафедра','ID працівника','Назва та реквізити публікації(посилання)','Назва наукометричної бази')
        self.bd = Side(style='thick', color="000000")
        self.bl = Side(style='thin', color="000000")

        self.wb : Workbook  = Workbook()
        self.ws : Worksheet = self.wb.active
        self.output: BytesIO = BytesIO()
 
    def set_Header(self,i_row,Doc_Name):
     for i_col,val in enumerate(Doc_Name):        
        adr=get_column_letter(i_col+1)+str(i_row)
        if val: self.ws[adr].value = val
        self.ws[adr].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        self.ws[adr].font = Font(bold=True)
        self.ws[adr].border = Border(left=self.bd, top=self.bd, right=self.bd, bottom=self.bd)
        self.ws[adr].fill = PatternFill('solid',fgColor='ffbf80')

    def set_Width(self,d):
        for i,val in enumerate(d):
            self.ws.column_dimensions[get_column_letter(i+1)].width = val

    def drawColorYellow(self,i):
        for tyty in self.ws['A'+ str(i)+':H'+ str(i)]:
            for strk in tyty:
                strk.fill = PatternFill('solid',fgColor='FFFF00')
#----------------------------------------------------------------

class ScopusExportExcel(ExportExcel):

    def write_full_aticl(self,total_list):
        self.ws.title ='ALL_Aticles'
        self.set_Width(self.mas_Doc)
        self.set_Header(1,self.all_title)
        fist_aticle='-1'
        i_index=1
        count=1
        summ_note = 0 
        for record in total_list:
            if fist_aticle != record[7]:   #('#','Название статьи',"Год","Тип публ.",'Авторы/Автор','Кафедра','цитирования')
                i_index+=1
                self.ws['A'+str(i_index)]=count
                self.ws['B'+str(i_index)]=record[0]
                self.ws['C'+str(i_index)]=record[4]
                self.ws['D'+str(i_index)]=record[8]
                self.ws['E'+str(i_index)]=record[1]
                if record[10].isdigit(): 
                    self.ws['G'+str(i_index)]=int(record[10])
                    summ_note += int(record[10])             
                else:
                    self.ws['G'+str(i_index)]=record[10]

                i_index+=1
                count+=1
                if (fist_aticle != record[7]) and (not record[2]) and (not record[3]):
                    self.drawColorYellow(i_index-1)
                fist_aticle=record[7]
            if not record[2]  and not record[3]:
                continue  
            self.ws['E'+str(i_index)]=record[2]
            self.ws['F'+str(i_index)]=record[3]
            i_index+=1
        self.ws['G'+str(i_index)] = summ_note


    def create_report_article(self, total_list):
        self.write_full_aticl(total_list)
        # output = BytesIO()
        self.wb.save(self.output)
        return self.output.getvalue()
        # return save_virtual_workbook(self.wb)  

    def create_report_author(self, total_list):
        pass

    def create_report_sum(self,total_list):
        self.ws.title ='Сумма цитирований по кафедрам'
        self.set_Width(self.mas_note)
        self.set_Header(1,self.title_note)
        fist_aticle='-1'
        i_index=1
        dep_ind=2
        count=0
        for record in total_list:
            if fist_aticle != record[1]:   #('#','Название статьи',"Год","Тип публ.",'Авторы/Автор','Кафедра')
                i_index+=1
                fist_aticle = record[1]
                self.ws['A'+str(i_index)]=record[1] if record[1]  else 'Не найдена'
                i_index+=1
                if i_index != 3:
                    self.ws['B'+str(dep_ind)]=count                    
                #    count=int(record[2])
                    dep_ind=i_index-1
                #else:
                    #i_index+=1                    
                count=int(record[2])
                self.ws['C'+str(i_index)]=record[0]
                self.ws['D'+str(i_index)]=record[2]

            else: 
                i_index+=1                    
                self.ws['C'+str(i_index)]=record[0]
                self.ws['D'+str(i_index)]=record[2]
                count+=int(record[2])
        
        self.ws['B'+str(dep_ind)]=count

        output = BytesIO()
        self.wb.save(output)
        return output.getvalue()                            
        # return save_virtual_workbook(self.wb) 
    
    def create_author_with_article(self,list_export,dd:DataScForm):
        def count_limit(dd:DataScForm,id):
            if not dd['sc_bool_limit']: return False
            count = 0
            for rec in list_export:
                if rec[0] == id: count+=1
            return count < dd['sc_input_limit']

            
        self.ws.title ='Авторы со статьями по кафедрам'
        self.set_Width(self.mas_aut_)
        self.set_Header(1,self.title_aut_)
        i_index=2
        fist_aut='-1'
        count_aut=1
        count=1
        id_Limit=None
        for record in list_export:
            if fist_aut != record[0]:   #('#','Название статьи',"Год","Тип публ.",'Авторы/Автор','Кафедра') 
                if id_Limit == record[0]: continue
                if count_limit(dd,record[0]):
                    id_Limit = record[0]
                    continue
                else: 
                    id_Limit=None
                # i_index+=2
                fist_aut = record[0]
                self.ws['A'+str(i_index)]=count_aut
                self.ws['B'+str(i_index)]=record[1]
                self.ws['C'+str(i_index)]=record[4]
                self.ws['D'+str(i_index)]=record[12]
                self.ws['F'+str(i_index)]="Scopus"


                count=1
                count_aut+=1
            # else: 
            str_to_write=""                    
            if record[8] :
                    str_to_write=f'{record[8]}'
            if record[9] :
                    str_to_write+=f'({record[9]})'
            if record[10]:
                if str_to_write :
                    str_to_write+=f':{record[10]};'
                else:
                    str_to_write+=f' p.{record[10]};'
            elif str_to_write:
                    str_to_write+=';'
            


            self.ws['E'+str(i_index)]=f'{count}. {record[6]}; {record[2]}; {record[7]}; {record[3]}; {str_to_write} {"DOI:" + record[11] if record[11] else ""}'
            # self.ws['D'+str(i_index)]=record[2]
            # self.ws['E'+str(i_index)]=record[3]
            i_index+=1
            count+=1       
        
        # output = BytesIO()
        self.wb.save(self.output)
        return self.output.getvalue()        
        # return save_virtual_workbook(self.wb) 
    

    def create_green_table(self,records):
        self.ws.title ='Экспорт_базы_данных'
        widthColumn=(8,35,20,14,8,11,9,8,11,9,7,12,7,18,9,7,22,100)
        list_range_merge=('A1:A2','B1:B2','C1:C2','D1:D2','E1:G1','H1:J1','K1:K2','L1:L2','M1:M2','N1:P1','Q1:Q2','R1:R2')
        title_1=('Ідентифікатор співробітника','Прізвище ім`я та по батькові','ORCID ID','ID Scopus','Scopus','','',
                 'Web of Science','','','ElАr KhNURE','Каф','Раб','Googlescholar','','','Профіль на NURE.UA','NameLat')
        title_2=('','','','','документи','цитування','індекс Гірша','документи','цитування','індекс Гірша',
                 '','','','ID','Кількість цитувань','H-Index','','')
             
        self.set_Width(widthColumn)
        for range_merge in list_range_merge:
            self.ws.merge_cells(range_merge)
        self.set_Header(1,title_1)
        self.set_Header(2,title_2)
        self.ws.freeze_panes = self.ws['A3']
        y_index=3
        for item in records:
            for x_index,data_cell in enumerate(item,1):
                x = x_index if x_index < 11 else  x_index + 1 if 11 <= x_index < 14  else x_index + 4
                address = get_column_letter(x)+str(y_index)
                
                if x_index == 3 and data_cell:
                    self.ws[address].hyperlink = f'https://orcid.org/{data_cell}'
                    self.ws[address].font= Font(color=Color(BLUE))
                elif x_index == 4 and  data_cell:
                    self.ws[address].hyperlink = f'https://www.scopus.com/authid/detail.uri?authorId={data_cell}'
                    self.ws[address].font= Font(color=Color(BLUE))
                elif x_index == 13 and  data_cell:
                    self.ws[address].hyperlink = f'https://scholar.google.com.ua/citations?user={data_cell}&hl=ru'
                    self.ws[address].font= Font(color=Color(BLUE))

                elif x_index == 12:
                    data_cell = 'Да' if data_cell else 'Нет'                

                self.ws[address] = data_cell

                # форматирование таблицы--------------------------------------------

                if 2 < x_index < 14 :
                    self.ws[address].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                else:
                    self.ws[address].alignment = Alignment(vertical='center',wrap_text=True)

                self.ws[address].border = Border(left=self.bl, top=self.bl, right=self.bl, bottom=self.bl)
                self.ws[address].fill = PatternFill('solid',fgColor='66FF66') 
            y_index+=1    

        # output = BytesIO()
        self.wb.save(self.output)
        return self.output.getvalue()

        # return save_virtual_workbook(self.wb) 
    
    def create_report_authors_with_stat(self,list_export,wos=False):
        self.ws.title ='Authors_with_statistics'
        widthColumn = (8,45,15,20,15,15,15)  if not wos else (8,45,15,20,15,15,15,15)
        list_range_merge=('A1:A2','B1:B2','C1:C2','D1:D2','E1:G1') if not wos else ('A1:A2','B1:B2','C1:C2','D1:D2','E1:E2','F1:H1')
        title_1=('№ п\п ','Прізвище ім`я та по батькові','Кафедра','ID Scopus', 'Scopus') if not wos else ('№ п\п ','Прізвище ім`я та по батькові','Кафедра','ID Orcid', 'Reseacher','WOS')
        title_2=('','','','','Документи','Цитування','H-Index') if not wos else ('','','','','','Документи','Цитування','H-Index')         
        self.set_Width(widthColumn)
        for range_merge in list_range_merge:
            self.ws.merge_cells(range_merge)
        self.set_Header(1,title_1)
        self.set_Header(2,title_2)
        y_index=3
        for item in list_export:
            for x_index,data_cell in enumerate(item,1):
                
                if x_index > 7 and not wos: 
                    continue                
                elif x_index > 8 and wos:
                    continue    
                address = get_column_letter(x_index)+str(y_index)                
                if x_index == 4 and data_cell:
                    self.ws[address].hyperlink = f'https://www.scopus.com/authid/detail.uri?authorId={data_cell}'
                    self.ws[address].font= Font(color=Color(BLUE))
                                
                self.ws[address] = data_cell if x_index != 1 else y_index - 2
                self.ws[address].border = Border(left=self.bl, top=self.bl, right=self.bl, bottom=self.bl)
                if  x_index != 2 :
                    self.ws[address].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                else:
                    self.ws[address].alignment = Alignment(vertical='center',wrap_text=True)

            y_index+=1  
        
        # output = BytesIO()
        self.wb.save(self.output)
        return self.output.getvalue()
